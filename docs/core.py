"""
Flight Realization Migrator — core logic
========================================
Logika inti migrasi realisasi penerbangan, **lepas dari GUI**.
"""

import re
import calendar
import datetime
import shutil
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Konstanta
# ---------------------------------------------------------------------------

WITA_AIRPORTS = {'DPS', 'UPG', 'LOP', 'BPN', 'AAP', 'BDJ', 'MDC', 'KDI', 'PLW', 'KOE', 'TRK', 'LBJ', 'BMU', 'TMC', 'MOF'}
WIT_AIRPORTS = {'AMQ', 'DJJ', 'SOQ', 'TIM', 'TTE', 'MKQ'}

def adjust_time_str(time_val, offset_hours):
    if time_val is None:
        return None
    time_str = str(time_val).strip()
    if not time_str or time_str.lower() in ('none', 'nan', 'nat'):
        return time_str
    
    m = re.search(r'(\d{1,2}):(\d{2})', time_str)
    if not m:
        if isinstance(time_val, (datetime.time, datetime.datetime)):
            time_str = time_val.strftime('%H:%M')
            m = re.search(r'(\d{1,2}):(\d{2})', time_str)
            if not m: return time_val
        else:
            return time_val
            
    h = int(m.group(1))
    m_val = int(m.group(2))
    
    h = (h + offset_hours) % 24
    if h < 0: h += 24
    
    new_time = f"{h:02d}:{m_val:02d}"
    return time_str.replace(m.group(0), new_time)

INDONESIAN_MONTHS = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

# Nama bulan Indonesia (lengkap & singkat) untuk mem-parse tanggal di PDF
MONTH_NAME_TO_NUM = {
    'januari': 1, 'februari': 2, 'maret': 3, 'april': 4, 'mei': 5, 'juni': 6,
    'juli': 7, 'agustus': 8, 'september': 9, 'oktober': 10, 'november': 11,
    'desember': 12,
}
MONTH_ABBR = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR", 5: "MEI", 6: "JUN",
    7: "JUL", 8: "AGU", 9: "SEP", 10: "OKT", 11: "NOV", 12: "DES",
}
MONTH_ABBR_TO_NUM = {v: k for k, v in MONTH_ABBR.items()}

# Pola parsing baris master, contoh: "QG726 -QG430 /CGK -SUB -BPN"
# Mendukung QG, ID, IW, IU
FLIGHT_PATTERN = re.compile(
    r'^\s*([A-Z0-9]+)\s*-\s*([A-Z0-9]+)\s*/\s*(\w+)\s*-\s*(\w+)\s*-\s*(\w+)\s*$'
)

# Pola tanggal di kolom master: "DD/MM" atau "DD/MM/YYYY"
DATE_PATTERN = re.compile(r'^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?')

def parse_excel_date(date_val):
    """
    Parse date_val dari sel Excel secara fleksibel:
    - Objek datetime.date / datetime.datetime
    - Objek pd.Timestamp
    - Teks dengan format 'DD/MM/YYYY', 'DD/MM', atau 'YYYY-MM-DD'
    """
    if pd.isna(date_val):
        return None, None, None
        
    if isinstance(date_val, (datetime.date, datetime.datetime)):
        return date_val.day, date_val.month, date_val.year
        
    if hasattr(date_val, 'to_pydatetime'): # pd.Timestamp
        dt = date_val.to_pydatetime()
        return dt.day, dt.month, dt.year
        
    val_str = str(date_val).strip()
    
    # Coba format YYYY-MM-DD (dengan atau tanpa jam)
    m_ymd = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', val_str)
    if m_ymd:
        return int(m_ymd.group(3)), int(m_ymd.group(2)), int(m_ymd.group(1))
        
    # Coba format DD/MM/YYYY
    m_dmy = DATE_PATTERN.match(val_str)
    if m_dmy:
        day = int(m_dmy.group(1))
        mon = int(m_dmy.group(2))
        year = None
        if m_dmy.group(3):
            y_str = m_dmy.group(3)
            if len(y_str) == 2:
                year = 2000 + int(y_str)
            else:
                year = int(y_str)
        return day, mon, year
        
    return None, None, None

# --- Pola untuk file PPRP (PDF) ---
# Nomor surat di kepala surat, mis. "Nomor : AU.012/44/7/DJPU-DAU-2026"
PPRP_NOMOR_PATTERN = re.compile(r'Nomor\s*:\s*(AU\.\S+?-\d{4})')
# Tanggal "DD NamaBulan YYYY" di dalam sel tanggal penerbangan
PPRP_DATE_PATTERN = re.compile(r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})')
# Jam "HH:MM"
TIME_PATTERN = re.compile(r'^(\d{1,2}):(\d{2})$')


def normalize_flight_key(s):
    """Normalisasi nomor penerbangan: 'QG-723' / 'QG723' -> 'QG723'."""
    return str(s).replace('-', '').replace(' ', '').upper()


def normalize_surat(no_surat):
    """Samakan gaya nomor surat dengan form: 'DJPU-DAU' -> 'DRJU-DAU'."""
    return no_surat.replace('DJPU-DAU', 'DRJU-DAU')


def format_periode(tanggal_text):
    """'17 Juni 2026 29 Juni 2026' -> '17 JUN 2026/29 JUN 2026'."""
    parts = []
    for d, mon, y in PPRP_DATE_PATTERN.findall(tanggal_text):
        num = MONTH_NAME_TO_NUM.get(mon.lower())
        if num is None:
            continue
        parts.append(f"{int(d)} {MONTH_ABBR[num]} {y}")
    return '/'.join(parts[:2])


def parse_periode_dates(periode_str):
    """'17 JUN 2026/29 JUN 2026' -> (datetime.date(2026, 6, 17), datetime.date(2026, 6, 29))"""
    if not periode_str or '/' not in periode_str:
        return None, None
    parts = periode_str.split('/')
    if len(parts) != 2:
        return None, None
    
    parsed_dates = []
    for p in parts:
        p = p.strip()
        sub_parts = p.split()
        if len(sub_parts) != 3:
            return None, None
        try:
            d = int(sub_parts[0])
            m = MONTH_ABBR_TO_NUM.get(sub_parts[1].upper())
            y = int(sub_parts[2])
            if m is None:
                return None, None
            parsed_dates.append(datetime.date(y, m, d))
        except Exception:
            return None, None
            
    return parsed_dates[0], parsed_dates[1]


# ---------------------------------------------------------------------------
# Parser PDF PPRP (Perubahan Penetapan Pelaksanaan Rute Penerbangan)
# ---------------------------------------------------------------------------

def parse_pprp_pdf(path):
    """Baca satu file PDF PPRP dan kembalikan daftar segmen perubahan."""
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError(
            "Membaca PDF memerlukan library 'pdfplumber'.\n"
            "Install: pip install pdfplumber"
        )

    segments = []
    with pdfplumber.open(path) as pdf:
        no_surat = None
        tgl_surat = None
        
        # Ambil nomor surat & tanggal surat dari halaman pertama
        for page in pdf.pages:
            txt = page.extract_text() or ''
            # Coba cari nomor dan tanggal dalam satu baris
            m = re.search(r'Nomor\s*:\s*(AU\.\S+?-\d{4})\s*(?:Jakarta|Tangerang|Banten)?,?\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})', txt)
            if m:
                no_surat = normalize_surat(m.group(1))
                tgl_surat = m.group(2).strip()
                break
            else:
                m2 = PPRP_NOMOR_PATTERN.search(txt)
                if m2:
                    no_surat = normalize_surat(m2.group(1))
                    # Coba cari tanggal yang ada di halaman pertama
                    m_tgl = re.search(r'(?:Jakarta|Tangerang|Banten)?,?\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})', txt)
                    if m_tgl:
                        tgl_surat = m_tgl.group(1).strip()
                    break

        # Cari halaman yang memuat tabel MENJADI dan parse tabelnya
        for page in pdf.pages:
            txt = (page.extract_text() or '').upper()
            if 'MENJADI' not in txt:
                continue
            for table in page.extract_tables():
                for row in table:
                    seg = _parse_pprp_row(row, no_surat)
                    if seg:
                        segments.append(seg)

    return {'no_surat': no_surat, 'tgl_surat': tgl_surat, 'segments': segments}


def _parse_pprp_row(row, no_surat):
    """Ubah satu baris tabel MENJADI menjadi segmen, atau None jika bukan data."""
    if not row:
        return None
    cells = [(c or '').replace('\n', ' ').strip() for c in row]

    # Cari nomor penerbangan (QG###, ID###, etc) di salah satu sel
    flight = None
    fi = None
    for i, c in enumerate(cells):
        if re.fullmatch(r'(?:QG|ID|IW|IU)\d+[A-Z]?', c):
            flight = c
            fi = i
            break
    if flight is None:
        return None

    # STD & STA = dua sel jam pertama setelah kolom nomor penerbangan
    times = [c for c in cells[fi + 1:] if TIME_PATTERN.match(c)]
    if len(times) < 2:
        return None
    etd = _to_time(times[0])
    eta = _to_time(times[1])

    # Hari operasi = angka 6-7 digit; periode = teks tanggal panjang
    hari = ''
    for c in cells[fi + 1:]:
        if re.fullmatch(r'\d{6,7}', c):
            hari = c
            break
    periode = ''
    for c in cells[fi + 1:]:
        if PPRP_DATE_PATTERN.search(c):
            periode = format_periode(c)
            break

    return {
        'flight': flight,
        'etd': etd,
        'eta': eta,
        'periode': periode,
        'hari': hari,
        'no_surat': no_surat,
    }


def _to_time(hhmm):
    h, m = TIME_PATTERN.match(hhmm).groups()
    return datetime.time(int(h) % 24, int(m))


def build_changes_map(pdf_paths, log=None):
    """Gabungkan semua PDF -> dict flight_key -> list segmen (urut sesuai input)."""
    changes = {}
    for p in pdf_paths:
        data = parse_pprp_pdf(p)
        if log:
            log(f"       {Path(p).name}: surat {data['no_surat']} ({data['tgl_surat']}), "
                f"{len(data['segments'])} baris MENJADI")
        for seg in data['segments']:
            seg['tgl_surat'] = data['tgl_surat']
            key = normalize_flight_key(seg['flight'])
            changes.setdefault(key, []).append(seg)
    return changes


# ---------------------------------------------------------------------------
# Parser PDF WTT (Working Time Table)
# ---------------------------------------------------------------------------

def build_wtt_map(pdf_paths, log=None):
    """Gabungkan semua WTT PDF -> dict flight_key -> jam keberangkatan (ETD)."""
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError(
            "Membaca PDF memerlukan library 'pdfplumber'.\n"
            "Install: pip install pdfplumber"
        )

    wtt_data = {}
    # Pola: 15:00 16:10 QG423
    pattern = re.compile(r'(\d{1,2}:\d{2})\s+(\d{1,2}:\d{2})\s+((?:QG|ID|IW|IU)\d+[A-Z]?)')
    
    for p in pdf_paths:
        with pdfplumber.open(p) as pdf:
            count = 0
            for page in pdf.pages:
                txt = page.extract_text() or ''
                for match in pattern.finditer(txt):
                    etd = match.group(1)
                    flt = normalize_flight_key(match.group(3))
                    wtt_data[flt] = etd
                    count += 1
            if log:
                log(f"       {Path(p).name}: {count} jadwal ATD (WTT) ditemukan")
                
    return wtt_data


# ---------------------------------------------------------------------------
# Kelas Utama Migrator
# ---------------------------------------------------------------------------

class Migrator:
    def __init__(self, master_path, form_path, output_path, month, year,
                 hub='SUB', pdf_paths=None, log=None, progress=None,
                 start_date=None, end_date=None, timezone_mode='default', wtt_paths=None):
        self.master_path = master_path
        self.form_path = form_path
        self.output_path = output_path
        self.month = month
        self.year = year
        self.hub = hub.upper()
        self.pdf_paths = list(pdf_paths or [])
        self.wtt_paths = list(wtt_paths or [])
        self._log = log or (lambda msg: None)
        self._progress = progress or (lambda pct: None)
        self.start_date = start_date
        self.end_date = end_date
        self.timezone_mode = timezone_mode.lower()
        self.cols = {}
        self.inserted_rows = set()

    def run(self):
        """Jalankan migrasi. Kembalikan dict ringkasan; raise bila gagal."""
        self._log(f"[1/6] Membaca master: {self.master_path}")
        master_df = self._read_master()
        self._log(f"       {len(master_df)} baris terbaca")
        self._progress(20)

        self._log(
            f"[2/6] Parsing keberangkatan dari {self.hub} "
            f"untuk {INDONESIAN_MONTHS[self.month]} {self.year}"
        )
        present, flight_atd = self._parse_master(master_df)
        self._log(f"       {len(present)} (flight, tanggal) ditemukan, {len(flight_atd)} flight ATD")
        self._progress(40)

        self._log(f"[3/6] Menyalin template ke output: {self.output_path}")
        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        if self.form_path != self.output_path:
            shutil.copy(self.form_path, self.output_path)
        wb = load_workbook(self.output_path)
        
        # Cari sheet CITILINK jika ada, jika tidak pakai yang aktif
        sheet_name = 'CITILINK'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._log(f"       Menggunakan sheet: '{sheet_name}'")
        else:
            ws = wb.active
            self._log(f"       Menggunakan active sheet: '{ws.title}'")
            
        self._progress(45)

        # Deteksi format template & pemetaan kolom
        self.cols = self._detect_and_map_columns(ws)
        self._log(f"       Format terdeteksi: {self.cols['format'].upper()} (Day 1 di kolom {self.cols['day_start']})")

        # Baca PDF PPRP (jika ada) & sisipkan blok "Perubahan"
        inserted = 0
        if self.pdf_paths:
            self._log(f"[4/6] Membaca {len(self.pdf_paths)} file PDF PPRP")
            changes = build_changes_map(self.pdf_paths, self._log)
            headers = self._find_flight_headers(ws)
            inserted = self._insert_perubahan_blocks(ws, headers, changes)
            self._log(f"       {inserted} blok perubahan disisipkan")
        else:
            self._log("[4/6] Tidak ada PDF PPRP (lewati penyisipan perubahan)")
        self._progress(60)

        # Baca PDF WTT jika ada
        wtt_data = {}
        if self.wtt_paths:
            self._log(f"[4.5/6] Membaca {len(self.wtt_paths)} file PDF WTT")
            wtt_data = build_wtt_map(self.wtt_paths, self._log)
        
        self._log("[5/6] Mencari blok flight di form")
        headers = self._find_flight_headers(ws)
        self._log(f"       {len(headers)} flight block ditemukan")
        if not headers:
            raise RuntimeError(
                "Tidak ada flight block ditemukan di form. "
                "Pastikan kolom B (flight number) berisi 'QG-XXX'."
            )
        self._progress(75)

        self._log("[6/6] Mengisi nilai per tanggal")
        results = self._fill_form(ws, headers, present, flight_atd, wtt_data)
        self._progress(90)

        # Update tanggal tanda tangan
        self._update_signature_date(ws)

        wb.save(self.output_path)
        self._log(f"\nSUCCESS: Selesai. File disimpan di:\n  {self.output_path}")
        self._progress(100)

        filled_rows = sum(1 for r in results if r.get('status') == 'ok')
        return {
            'results': results,
            'output': self.output_path,
            'inserted': inserted,
            'inserted_blocks': inserted,
            'filled_rows': filled_rows,
        }

    # ------------------------------------------------------------------ IO

    def _read_master(self):
        p = Path(self.master_path)
        suffix = p.suffix.lower()
        if suffix == '.xls':
            try:
                return pd.read_excel(self.master_path, header=None, engine='xlrd')
            except ImportError:
                raise RuntimeError(
                    "File .xls memerlukan library 'xlrd'.\n"
                    "Install: pip install xlrd\n"
                    "Atau: konversi file ke .xlsx terlebih dahulu."
                )
        elif suffix in ('.xlsx', '.xlsm'):
            return pd.read_excel(self.master_path, header=None, engine='openpyxl')
        else:
            raise RuntimeError(f"Format file master tidak didukung: {suffix}")

    # ------------------------------------------------------------- Parsing

    def _parse_master(self, df):
        """Kembalikan set of (flight_key, day) untuk bulan/tahun yang dipilih dan mapping ATD."""
        present = set()
        atd_map = {}
        num_days = calendar.monthrange(self.year, self.month)[1]
        
        # Cari index kolom ATD secara dinamis (biasanya 9, tapi bisa bergeser)
        atd_col_idx = 9
        for r_idx in range(min(5, len(df))):
            for c_idx, val in enumerate(df.iloc[r_idx].values):
                if isinstance(val, str) and 'Actual Time of Departure' in val:
                    atd_col_idx = c_idx
                    break
            else:
                continue
            break

        for _, row in df.iterrows():
            if len(row) < 2:
                continue
            date_val = row.iloc[0]
            flt_val = row.iloc[1]
            if pd.isna(date_val) or pd.isna(flt_val):
                continue

            m = FLIGHT_PATTERN.match(str(flt_val))
            if not m:
                continue
            arr_flt, dep_flt, org, via, dst = m.groups()

            # Hanya keberangkatan dari hub
            if via.upper() != self.hub:
                continue

            day, mon, yr = parse_excel_date(date_val)
            if mon is None or day is None:
                continue
            if mon != self.month or day < 1 or day > num_days:
                continue

            flight_key = normalize_flight_key(dep_flt)
            present.add((flight_key, day))
            
            # Ekstrak ATD dari kolom dinamis
            if len(row) > atd_col_idx:
                atd_val = row.iloc[atd_col_idx]
                if pd.notna(atd_val):
                    atd_str = str(atd_val).strip()
                    if ':' in atd_str:
                        # Parsing format "HH:MM", "H:MM", "HH:MM:SS"
                        parts = atd_str.split(':')
                        if len(parts) >= 2:
                            # H:MM -> 0H:MM
                            h = parts[0][-2:].zfill(2)
                            m_minute = parts[1][:2].zfill(2)
                            atd_str = f"{h}:{m_minute}"
                    atd_map.setdefault(flight_key, []).append(atd_str)

        # Cari ATD terbanyak (modus) untuk tiap flight
        flight_atd = {}
        for flt, atds in atd_map.items():
            if atds:
                # Mengambil mode dengan penanganan tie-breaker (pilih yang pertama disort)
                flight_atd[flt] = max(sorted(set(atds)), key=atds.count)

        return present, flight_atd

    # -------------------------------------------------------------- Form

    def _detect_and_map_columns(self, ws):
        """Mendeteksi format template (New Side-by-Side S26 atau Old Block-based)."""
        is_new_format = False
        month_name = INDONESIAN_MONTHS[self.month].upper()
        month_col = None
        
        # Cari nama bulan di Baris 2 untuk mendeteksi format baru
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=2, column=c).value
            if isinstance(v, str) and v.upper().strip() in ('MARET', 'APRIL', 'MEI', 'JUNI', 'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER', 'JANUARI', 'FEBRUARI'):
                is_new_format = True
            if isinstance(v, str) and v.upper().strip() == month_name:
                month_col = c

        if is_new_format:
            self.month_name = month_name
            if not month_col:
                month_col = 13
                for c in range(1, 500):
                    v2 = ws.cell(2, c).value
                    v3 = ws.cell(3, c).value
                    v4 = ws.cell(4, c).value
                    if v4 == 1 and (month_name.lower() in str(v2).lower() or month_name.lower() in str(v3).lower()):
                        month_col = c
                        break
            
            # Default mapping awal untuk format baru
            cols = {
                'no': 1,
                'flight': 2,
                'to': 3,
                'etd': 4,
                'eta': 5,
                'atd': 6,
                'periode': 7,
                'surat': 8,
                'date': 9,
                'hari': 10,
                'day_start': month_col,
                'day_end': month_col + calendar.monthrange(self.year, self.month)[1] - 1,
                'total': month_col - 1,
                'block_merge_cols': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
                'format': 'new'
            }

            # Cari letak kolom metadata sebelum month_col secara mundur
            for c in range(month_col - 1, max(1, month_col - 16), -1):
                v2 = ws.cell(row=2, column=c).value
                v4 = ws.cell(row=4, column=c).value
                v2_str = str(v2).upper().strip() if v2 else ""
                v4_str = str(v4).upper().strip() if v4 else ""
                
                if v2_str == 'SURAT':
                    cols['surat'] = c
                    cols['date'] = c + 1
                elif v2_str == 'DAY OF FLIGHT':
                    cols['hari'] = c
                elif v2_str == 'FLIGHT NUMBER' or v4_str in ('FLT NO', 'FLT NO.'):
                    cols['flight'] = c
                elif v4_str == 'TO':
                    cols['to'] = c
                elif v2_str == 'SCHEDULE' or v4_str == 'STD':
                    cols['etd'] = c
                    cols['eta'] = c + 1
                elif v2_str == 'ACTUAL' or v4_str == 'ATD':
                    cols['atd'] = c
                elif v2_str == 'PERIODE':
                    cols['periode'] = c
                elif v2_str == 'REALISASI':
                    cols['total'] = c
                elif (v2_str == 'NO' or v4_str == 'NO') and v2_str != 'SURAT':
                    cols['no'] = c

            merge_cols = [
                cols['no'], cols['flight'], cols['to'], cols['etd'], cols['eta'],
                cols['periode'], cols['surat'], cols['date'], cols['hari']
            ]
            if 'atd' in cols and cols['atd'] is not None:
                merge_cols.append(cols['atd'])
                
            cols['block_merge_cols'] = sorted(list(set(merge_cols)))
            return cols
        else:
            # Format lama
            return {
                'no': 1,
                'flight': 2,
                'to': 3,
                'etd': 4,
                'eta': 5,
                'atd': 6,
                'periode': 8,
                'surat': 9,
                'date': 10,  # COL_TIPE
                'hari': 11,
                'day_start': 14,
                'day_end': 44,
                'total': 45,
                'block_merge_cols': [1, 2, 3, 4, 5, 6, 8, 9, 10, 11, 12],
                'format': 'old'
            }

    def _find_flight_headers(self, ws):
        """Cari semua baris yang berisi flight number di kolom flight."""
        headers = []
        col_flight = self.cols['flight'] if self.cols else 2
        last_flight = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, col_flight).value
            v_fr = ws.cell(r, 174).value
            flight_str = None
            if v is not None and str(v).strip().upper().startswith('QG'):
                flight_str = str(v).strip()
            elif v_fr is not None and str(v_fr).strip().upper().startswith('QG'):
                flight_str = str(v_fr).strip()
            if flight_str is None:
                last_flight = None
                continue
            s = flight_str
            # Mendukung QG, ID, IW, IU dengan/tanpa dash
            if re.match(r'^(?:QG|ID|IW|IU)-?\d+[A-Z]?$', s):
                norm_name = s.replace('-', '').replace(' ', '').upper()
                if norm_name != last_flight:
                    headers.append((r, s))
                    last_flight = norm_name
            else:
                last_flight = None
        return headers

    # ------------------------------------------- Penyisipan blok "Perubahan"

    def _existing_perubahan_keys(self, ws, headers):
        """Set (flight, periode) untuk blok Perubahan yang sudah ada."""
        keys = set()
        
        # Temukan semua kolom 'periode' di Row 2 jika format baru
        periode_cols = []
        if self.cols['format'] == 'new':
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=2, column=c).value
                if isinstance(v, str) and v.upper().strip() == 'PERIODE':
                    periode_cols.append(c)
        else:
            periode_cols = [self.cols['periode']]

        for i, (r, name) in enumerate(headers):
            end_r = headers[i + 1][0] - 1 if i + 1 < len(headers) else min(r + 20, ws.max_row)
            for r_idx in range(r, end_r + 1):
                is_perubahan = False
                if self.cols['format'] == 'new':
                    is_perubahan = True  # Kumpulkan semua periode dari seluruh baris
                else:
                    tipe = ws.cell(r_idx, self.cols['date']).value
                    is_perubahan = tipe and 'perubahan' in str(tipe).lower()

                if is_perubahan:
                    for col in periode_cols:
                        val = ws.cell(r_idx, col).value
                        if val:
                            keys.add((
                                normalize_flight_key(name),
                                str(val).strip(),
                            ))
        return keys

    def _block_end(self, ws, start):
        """Batas bawah blok = akhir merge vertikal kolom A yang mulai di `start`."""
        if self.cols['format'] == 'new':
            return start
        for rng in ws.merged_cells.ranges:
            if rng.min_col == 1 and rng.max_col == 1 and rng.min_row == start:
                return rng.max_row
        return start

    def _unmerge_region(self, ws, top, bottom):
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= top and rng.max_row <= bottom:
                ws.unmerge_cells(str(rng))

    def _copy_row(self, ws, src, dst):
        """Salin nilai + style + tinggi baris dari baris src ke dst."""
        for col in range(1, ws.max_column + 1):
            s = ws.cell(src, col)
            d = ws.cell(dst, col)
            d.value = s.value
            if s.has_style:
                d.font = copy(s.font)
                d.border = copy(s.border)
                d.fill = copy(s.fill)
                d.number_format = s.number_format
                d.protection = copy(s.protection)
                d.alignment = copy(s.alignment)
        if ws.row_dimensions[src].height is not None:
            ws.row_dimensions[dst].height = ws.row_dimensions[src].height

    def _shift_merged_ranges_below(self, ws, row_idx, amount=1):
        """Menggeser koordinat merged cell yang berada di bawah row_idx agar sinkron dengan baris yang bergeser."""
        to_remove = []
        to_add = []
        for rng in list(ws.merged_cells.ranges):
            # Jika seluruh range merged cell berada di bawah baris penyisipan
            if rng.min_row >= row_idx:
                to_remove.append(rng)
                to_add.append((rng.min_row + amount, rng.min_col, rng.max_row + amount, rng.max_col))
            # Jika range terpotong (min_row < row_idx <= max_row)
            elif rng.min_row < row_idx <= rng.max_row:
                to_remove.append(rng)
                # Bagian atas tetap di tempat
                if rng.min_row < row_idx - 1:
                    to_add.append((rng.min_row, rng.min_col, row_idx - 1, rng.max_col))
                # Bagian bawah bergeser ke bawah sebanyak amount
                if row_idx < rng.max_row + amount:
                    to_add.append((row_idx + amount, rng.min_col, rng.max_row + amount, rng.max_col))
                    
        for rng in to_remove:
            ws.merged_cells.ranges.remove(rng)
        for start_row, start_col, end_row, end_col in to_add:
            ws.merge_cells(start_row=start_row, start_column=start_col,
                           end_row=end_row, end_column=end_col)

    def _insert_perubahan_blocks(self, ws, headers, changes):
        """Sisipkan satu blok 'Perubahan' di bawah tiap flight yang berubah."""
        if not changes:
            return 0

        existing = self._existing_perubahan_keys(ws, headers)
        extents = {start: self._block_end(ws, start) for (start, _) in headers}
        first_start = headers[0][0]
        last_start = headers[-1][0]
        last_size = extents[last_start] - last_start + 1

        # Hapus merge di area blok hanya jika format lama
        if self.cols['format'] == 'old':
            self._unmerge_region(ws, first_start, extents[last_start])

        inserted = 0
        for start, name in reversed(headers):
            key = normalize_flight_key(name)
            segs = changes.get(key)
            if not segs:
                continue
            end = extents[start]
            blen = end - start + 1
            for seg in reversed(segs):
                sig = (key, seg['periode'])
                if sig in existing:
                    self._log(
                        f"       - {name}: blok '{seg['periode']}' sudah ada, dilewati"
                    )
                    continue
                ws.insert_rows(end + 1, blen)
                self._shift_merged_ranges_below(ws, end + 1, blen)
                for off in range(blen):
                    self._copy_row(ws, start + off, end + 1 + off)
                top = end + 1
                
                # Kosongkan kolom NO urut
                for off in range(blen):
                    ws.cell(top + off, self.cols['no']).value = None
                
                # Jika format baru, kosongkan kolom hari-hari (realisasi) semua bulan agar tidak terduplikasi
                if self.cols['format'] == 'new':
                    realization_cols = set()
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=2, column=c).value
                        if isinstance(v, str) and v.upper().strip() in (
                            'MARET', 'APRIL', 'MEI', 'JUNI', 'JULI', 'AGUSTUS', 
                            'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER', 'JANUARI', 'FEBRUARI'
                        ):
                            m_idx = MONTH_NAME_TO_NUM[v.lower().strip()]
                            days_in_month = calendar.monthrange(self.year, m_idx)[1]
                            realization_cols.update(range(c, c + days_in_month))
                            
                    for off in range(blen):
                        for c in realization_cols:
                            ws.cell(top + off, c).value = None
                
                # Tulis data perubahan
                ws.cell(top, self.cols['etd']).value = seg['etd']
                ws.cell(top, self.cols['eta']).value = seg['eta']
                ws.cell(top, self.cols['periode']).value = seg['periode']
                
                # Tulis no_surat dari file PPRP langsung ke baris perubahan (baris ke-2)
                # Baris asli (baris ke-1) tidak diubah
                if seg.get('no_surat'):
                    ws.cell(top, self.cols['surat']).value = seg['no_surat']
                    
                if seg['hari']:
                    ws.cell(top, self.cols['hari']).value = seg['hari']
                    
                self.inserted_rows.add(top)
                existing.add(sig)
                inserted += 1
                self._log(
                    f"       + {name}: blok Perubahan '{seg['periode']}' "
                    f"(ETD {seg['etd']:%H:%M} / ETA {seg['eta']:%H:%M})"
                )

        self._rebuild_blocks(ws, last_size)
        return inserted

    def _rebuild_blocks(self, ws, last_size):
        """Bangun ulang merge vertikal, penomoran kolom A, dan rumus total."""
        headers = self._find_flight_headers(ws)
        for i, (start, name) in enumerate(headers):
            end = headers[i + 1][0] - 1 if i + 1 < len(headers) else start + last_size - 1
            
            # Hanya merge vertikal jika format lama dan end > start
            if self.cols['format'] == 'old' and end > start:
                for col in self.cols['block_merge_cols']:
                    ws.merge_cells(start_row=start, start_column=col,
                                   end_row=end, end_column=col)
                                   
            ws.cell(start, self.cols['no']).value = i + 1
            self._rebuild_totals(ws, start, end)

    def _rebuild_totals(self, ws, start, end):
        """Perbaiki nomor baris pada rumus =COUNTIF(...) atau =SUM(...) tanpa mengubah kolomnya."""
        if not self.cols['total']:
            return
            
        for r in range(start, end + 1):
            total_cell = ws.cell(r, self.cols['total'])
            if type(total_cell).__name__ == 'MergedCell':
                continue
                
            from openpyxl.utils import get_column_letter
            c_start = get_column_letter(self.cols['day_start'])
            c_end = get_column_letter(self.cols['day_end'])
            
            # Paksa penulisan rumus untuk setiap baris di blok penerbangan ini
            if self.cols['format'] == 'new':
                total_cell.value = f'=COUNTIF({c_start}{r}:{c_end}{r}, "O")'
            else:
                total_cell.value = f'=SUM({c_start}{r}:{c_end}{r})'

    def _find_month_row(self, ws, start_row, end_row):
        """Cari baris di dalam blok yang cocok dengan bulan/tahun target."""
        if self.cols['format'] == 'new':
            return start_row

        yr_short = str(self.year)[-2:]
        month_name = INDONESIAN_MONTHS[self.month]
        month_short = month_name[:3].lower()
        col_month = 13  # M di format lama
        
        for r in range(start_row, end_row + 1):
            v = ws.cell(r, col_month).value
            if isinstance(v, (datetime.datetime, datetime.date)):
                if v.year == self.year and v.month == self.month:
                    return r
            elif isinstance(v, str):
                s = v.strip().lower()
                if s.startswith(month_short) and yr_short in s:
                    return r
        return None

    def _fill_form(self, ws, headers, present, flight_atd=None, wtt_data=None):
        results = []
        num_days = calendar.monthrange(self.year, self.month)[1]
        flight_atd = flight_atd or {}
        wtt_data = wtt_data or {}

        for i, (r, name) in enumerate(headers):
            flt = name
            flt_norm = normalize_flight_key(flt)
            
            if i + 1 < len(headers):
                end_r = headers[i + 1][0] - 1
            else:
                if self.cols['format'] == 'old':
                    end_r = self._block_end(ws, r)
                else:
                    end_r = r
                    for curr_r in range(r + 1, ws.max_row + 1):
                        v = ws.cell(curr_r, self.cols['flight']).value
                        if not v or str(v).replace('-', '').replace(' ', '').upper() != flt_norm:
                            break
                        end_r = curr_r
            
            # Konversi Zona Waktu pada blok ini (STD, STA, ATD lama)
            if self.timezone_mode != 'default':
                for r_idx in range(r, end_r + 1):
                    last_dest = ""
                    if self.cols.get('to'):
                        dest_val = ws.cell(r_idx, self.cols['to']).value
                        if dest_val: 
                            last_dest = str(dest_val).strip().upper()
                        
                    sta_offset = 7
                    if last_dest in WITA_AIRPORTS: sta_offset = 8
                    elif last_dest in WIT_AIRPORTS: sta_offset = 9
                    
                    if "QG484" in flt_norm:
                        # Aturan Khusus QG 484: Skip konversi STD/STA, hanya ATD yang dikonversi
                        std_diff = 0
                        sta_diff = 0
                        atd_diff = 0 if self.timezone_mode == 'utc' else 7
                    else:
                        if r_idx in self.inserted_rows:
                            # Baru disisipkan dari PPRP (PPRP murni Local Time)
                            std_diff = -7 if self.timezone_mode == 'utc' else 0
                            sta_diff = -sta_offset if self.timezone_mode == 'utc' else 0
                            atd_diff = -7 if self.timezone_mode == 'utc' else 0
                        else:
                            # Baris template lama (Template Asli adalah UTC)
                            std_diff = 0 if self.timezone_mode == 'utc' else 7
                            sta_diff = 0 if self.timezone_mode == 'utc' else sta_offset
                            atd_diff = 0 if self.timezone_mode == 'utc' else 7
                        
                    if self.cols.get('etd'):
                        c_cell = ws.cell(r_idx, self.cols['etd'])
                        if type(c_cell).__name__ != 'MergedCell':
                            c_val = c_cell.value
                            c_cell.value = adjust_time_str(c_val, std_diff)
                    if self.cols.get('eta'):
                        c_cell = ws.cell(r_idx, self.cols['eta'])
                        if type(c_cell).__name__ != 'MergedCell':
                            c_val = c_cell.value
                            c_cell.value = adjust_time_str(c_val, sta_diff)
                    if self.cols.get('atd'):
                        c_cell = ws.cell(r_idx, self.cols['atd'])
                        if type(c_cell).__name__ != 'MergedCell':
                            c_val = c_cell.value
                            c_cell.value = adjust_time_str(c_val, atd_diff)

            # Lewati isi kehadiran jika flight kosong untuk format lama
            month_row_old = None
            if self.cols['format'] == 'old':
                month_row_old = self._find_month_row(ws, r, end_r)
                if month_row_old is None:
                    self._log(
                        f"       ! {flt}: baris {INDONESIAN_MONTHS[self.month]}-{self.year} "
                        f"tidak ditemukan, dilewati"
                    )
                    results.append({
                        'flight': flt, 'row': None, 'total': None,
                        'status': f'row {INDONESIAN_MONTHS[self.month]}-{self.year} not found'
                    })
                    continue

            key = flt.replace('-', '').replace(' ', '').upper()

            cnt = 0
            last_filled_row = None
            for day in range(1, num_days + 1):
                is_pres = (key, day) in present
                curr_date = datetime.date(self.year, self.month, day)
                
                if self.cols['format'] == 'new':
                    val = 'O' if is_pres else 'X'
                else:
                    val = 1 if is_pres else 0
                    
                if self.cols['format'] == 'new':
                    target_row = r
                    best_start_dt = None
                    
                    for r_idx in range(r, end_r + 1):
                        periode_val = None
                        for col_idx in range(1, ws.max_column + 1):
                            cell_val = ws.cell(r_idx, col_idx).value
                            if isinstance(cell_val, str) and '/' in cell_val:
                                start_dt, end_dt = parse_periode_dates(cell_val)
                                if start_dt and end_dt:
                                    periode_val = cell_val
                                    break
                        
                        if periode_val:
                            start_dt, end_dt = parse_periode_dates(periode_val)
                            if start_dt <= curr_date <= end_dt:
                                if best_start_dt is None or start_dt > best_start_dt:
                                    best_start_dt = start_dt
                                    target_row = r_idx
                else:
                    target_row = month_row_old
                
                # Cek filter rentang tanggal
                in_range = True
                if self.start_date and curr_date < self.start_date:
                    in_range = False
                if self.end_date and curr_date > self.end_date:
                    in_range = False
                    
                target_cell = ws.cell(target_row, self.cols['day_start'] - 1 + day)

                # Cek apakah sel diblock warna (misal merah untuk summer season) khusus di bulan Maret
                is_blocked = False
                if self.month == 3:
                    if hasattr(target_cell, 'fill') and hasattr(target_cell.fill, 'fgColor'):
                        color = target_cell.fill.fgColor
                        if color.type == 'rgb' and color.rgb in ('FFFF0000', 'FF0000'):
                            is_blocked = True
                
                if is_blocked:
                    continue

                if in_range:
                    target_cell.value = val
                    if target_row is not None:
                        last_filled_row = target_row
                else:
                    target_cell.value = None
                if is_pres:
                    cnt += 1

            # Tulis ATD untuk bulan yang aktif (Timpa nilai lama jika ada)
            if self.cols.get('atd'):
                atd_val = None
                if flt_norm in wtt_data:
                    atd_val = wtt_data[flt_norm]
                    if self.timezone_mode == 'utc':
                        atd_val = adjust_time_str(atd_val, -7)
                elif flt_norm in flight_atd:
                    atd_val = flight_atd[flt_norm]
                    if self.timezone_mode == 'utc':
                        atd_val = adjust_time_str(atd_val, -7)
                
                if atd_val is not None:
                    t_row = target_row if target_row is not None else (month_row_old if month_row_old is not None else r)
                    atd_target_cell = ws.cell(t_row, self.cols['atd'])
                    if type(atd_target_cell).__name__ == 'MergedCell':
                        atd_target_cell = ws.cell(r, self.cols['atd'])
                    atd_target_cell.value = atd_val

            # Kosongkan kolom tanggal 29-31 kalau bulan pendek untuk semua baris di blok ini
            for r_idx in range(r, end_r + 1):
                if self.cols['format'] == 'new' or r_idx == month_row_old:
                    for day in range(num_days + 1, 32):
                        ws.cell(r_idx, self.cols['day_start'] - 1 + day).value = None

            # Tulis rumus total untuk semua baris di blok ini
            for r_idx in range(r, end_r + 1):
                if self.cols['format'] == 'new' or r_idx == month_row_old:
                    if self.cols['total']:
                        total_cell = ws.cell(r_idx, self.cols['total'])
                        from openpyxl.utils import get_column_letter
                        c_start = get_column_letter(self.cols['day_start'])
                        c_end = get_column_letter(self.cols['day_end'])
                        
                        if type(total_cell).__name__ == 'MergedCell':
                            pass
                        elif self.cols['format'] == 'new':
                            total_cell.value = f'=COUNTIF({c_start}{r_idx}:{c_end}{r_idx}, "O")'
                        else:
                            total_cell.value = f'=SUM({c_start}{r_idx}:{c_end}{r_idx})'

            log_row = last_filled_row if last_filled_row is not None else r
            self._log(f"       [OK] {flt}: {cnt}/{num_days} hari (baris {log_row})")
            results.append({
                'flight': flt, 'row': log_row, 'total': cnt, 'status': 'ok'
            })

        return results

    def _update_signature_date(self, ws):
        """Mencari teks pengesahan di pojok bawah (misal 'Sidoarjo,...') dan memperbaruinya."""
        import datetime
        import re
        
        now = datetime.datetime.now()
        months_id = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", 
                     "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
        current_month_str = months_id[now.month - 1]
        current_year = now.year
        
        # Scan maksimal 60 baris terbawah
        max_r = ws.max_row
        start_r = max(1, max_r - 60)
        
        # Pola: Menangkap kata berakhiran koma (misal: Sidoarjo,) diikuti opsional tanggal, lalu bulan & tahun
        # Contoh match: "Sidoarjo, 10 Maret 2026", "Surabaya, Agustus 2025"
        pattern = re.compile(r"([A-Za-z]+,)\s*(?:\d{1,2}\s+)?([A-Za-z]+)\s+(\d{4})", re.IGNORECASE)
        
        for r in range(start_r, max_r + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if type(cell).__name__ != 'MergedCell' and cell.value and isinstance(cell.value, str):
                    m = pattern.search(cell.value)
                    if m:
                        lokasi = m.group(1) # contoh: "Sidoarjo,"
                        # Menimpa dengan format tanpa tanggal: "Sidoarjo, Agustus 2026"
                        new_val = f"{lokasi} {current_month_str} {current_year}"
                        cell.value = pattern.sub(new_val, cell.value)
                        self._log(f"       [OK] Tanda tangan diperbarui: '{new_val}' pada sel baris {r}")
                        return # Selesai setelah menemukan 1 tanda tangan


# ---------------------------------------------------------------------------
# Fungsi pembungkus praktis
# ---------------------------------------------------------------------------

def migrate(master_path, form_path, output_path, month, year,
            hub='SUB', pdf_paths=None, log=None, progress=None,
            start_date=None, end_date=None, timezone_mode='default', wtt_paths=None):
    """Jalankan migrasi dan kembalikan ringkasan (termasuk daftar log)."""
    logs = []

    def _collect(msg):
        logs.append(msg)
        if log:
            log(msg)

    migrator = Migrator(
        master_path=master_path,
        form_path=form_path,
        output_path=output_path,
        month=month,
        year=year,
        hub=hub,
        pdf_paths=pdf_paths,
        log=_collect,
        progress=progress,
        start_date=start_date,
        end_date=end_date,
        timezone_mode=timezone_mode,
        wtt_paths=wtt_paths
    )
    summary = migrator.run()
    summary['log'] = logs
    return summary


def detect_master_month_and_year(master_path):
    """Membaca file master untuk mendeteksi bulan dan tahun mayoritas (paling sering muncul)."""
    p = Path(master_path)
    suffix = p.suffix.lower()
    try:
        if suffix == '.xls':
            df = pd.read_excel(master_path, header=None, engine='xlrd')
        elif suffix in ('.xlsx', '.xlsm'):
            df = pd.read_excel(master_path, header=None, engine='openpyxl')
        else:
            return None, None
            
        from collections import Counter
        month_counter = Counter()
        year_counter = Counter()
        
        # Ambil sampel hingga 500 baris pertama untuk performa cepat dan akurat
        sample_df = df.head(500)
        for _, row in sample_df.iterrows():
            if len(row) < 1:
                continue
            date_val = row.iloc[0]
            day, mon, yr = parse_excel_date(date_val)
            if mon is not None:
                month_counter[mon] += 1
                if yr is not None:
                    year_counter[yr] += 1
                    
        if month_counter:
            most_common_month = month_counter.most_common(1)[0][0]
            most_common_year = year_counter.most_common(1)[0][0] if year_counter else datetime.datetime.now().year
            return most_common_month, most_common_year
    except Exception:
        pass
    return None, None


def migrate_multi(master_paths, form_path, output_path, hub="SUB", pdf_paths=None, log=None, timezone_mode='default', wtt_paths=None):
    """Proses migrasi dari beberapa file master sekaligus ke dalam satu file template realisasi."""
    logs = []

    def _collect(msg):
        logs.append(msg)
        if log:
            log(msg)

    if not master_paths:
        raise ValueError("Daftar file master kosong.")

    _collect(f"=== Memulai Migrasi Multi-Bulan (Total: {len(master_paths)} file master) ===")
    
    # 1. Deteksi bulan dan tahun untuk masing-masing master file
    jobs = []
    for path in master_paths:
        mon, yr = detect_master_month_and_year(path)
        if mon is None:
            _collect(f"⚠️ Gagal mendeteksi bulan untuk file: {Path(path).name}. Dilewati.")
            continue
        jobs.append({
            'path': path,
            'month': mon,
            'year': yr
        })
        _collect(f"   - File: {Path(path).name} -> Terdeteksi Bulan: {mon}, Tahun: {yr}")
        
    if not jobs:
        raise ValueError("Tidak ada file master valid yang berhasil dideteksi bulannya.")
        
    # 2. Urutkan berdasarkan kronologi (tahun, lalu bulan)
    jobs.sort(key=lambda x: (x['year'], x['month']))
    
    # 3. Salin template ke output untuk inisialisasi
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(form_path, output_path)
    
    # 4. Jalankan migrasi berurutan pada file output yang sama
    summary_list = []
    current_input = output_path
    
    for i, job in enumerate(jobs):
        _collect(f"\n[Multi-Step {i+1}/{len(jobs)}] Memproses Bulan {job['month']} Tahun {job['year']}")
        summary = migrate(
            master_path=job['path'],
            form_path=current_input,
            output_path=output_path,
            month=job['month'],
            year=job['year'],
            hub=hub,
            pdf_paths=pdf_paths,
            log=_collect,
            timezone_mode=timezone_mode,
            wtt_paths=wtt_paths
        )
        summary_list.append(summary)
        # Langkah berikutnya akan membaca dari output_path yang baru saja ditulis
        current_input = output_path
        
    # Gabungkan summary
    total_filled = sum(s['filled_rows'] for s in summary_list)
    total_inserted = sum(s['inserted_blocks'] for s in summary_list)
    
    _collect(f"\n=== Selesai Migrasi Multi-Bulan ===")
    _collect(f"Status: Sukses")
    _collect(f"Total baris terisi: {total_filled}")
    _collect(f"Total blok perubahan disisipkan: {total_inserted}")
    
    return {
        'status': 'success',
        'filled_rows': total_filled,
        'inserted_blocks': total_inserted,
        'output': output_path,
        'details': summary_list,
        'log': logs
    }
