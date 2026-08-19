"""
Report Generator — Block-Format Excel Injection
================================================
Menginjeksi data dari database ScheduleVersion ke dalam template Excel otoritas
bandara (format blok musiman seperti form_realisasi_winter26.xlsx).

Algoritma:
1. Buka template Excel menggunakan openpyxl (load_workbook, keep_vba=False).
2. Deteksi otomatis koordinat kolom dari header (baris 5-6).
3. Temukan semua "blok flight" dari kolom B (flight number QG-xxx).
4. Untuk setiap flight yang ada di DB:
   a. Jika ada data PPRP -> duplikasi blok ke bawah, isi kolom Periode & Surat baru.
   b. Isi kolom absensi harian (1/0) per bulan dari data GHP (operational_flag).
5. Rebuild nomor urut (kolom A) dan perbaiki rumus SUM di kolom Total Realisasi.
6. Simpan sebagai file baru, kembalikan path-nya.
"""

import re
import calendar
import datetime
import os
import tempfile
from copy import copy
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator

from core.models import Project, ScheduleVersion


# ---------------------------------------------------------------------------
# Konstanta
# ---------------------------------------------------------------------------

INDONESIAN_MONTHS = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}
MONTH_ABBR = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR", 5: "MEI", 6: "JUN",
    7: "JUL", 8: "AGU", 9: "SEP", 10: "OKT", 11: "NOV", 12: "DES",
}


def _normalize_flight(s):
    """'QG-488' atau 'QG 488' -> 'QG488'."""
    return str(s).replace('-', '').replace(' ', '').upper().strip()


# ---------------------------------------------------------------------------
# Deteksi Kolom Template
# ---------------------------------------------------------------------------

def _detect_columns(ws):
    """
    Deteksi posisi kolom dari template berdasarkan header baris 5-6.
    Kembalikan dict dengan key nama kolom -> nomor kolom (1-indexed).
    Juga deteksi mapping bulan -> kolom_mulai (untuk pengisian harian).
    """
    cols = {
        'no': None,
        'flight': 2,
        'to': 3,
        'etd': 4,
        'eta': 5,
        'atd': 6,
        'ata': 7,
        'periode': 8,
        'surat': 9,
        'tipe': None,
        'hari': 10,
        'bulan_tahun': 13,
        'day_start': None,
        'total': None,
    }
    month_col_map = {}  # {month_num: start_col}

    # Scan baris 2 untuk menemukan nama bulan (kolom harian)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if not v:
            continue
        v_str = str(v).strip().upper()
        for m_num, m_name in INDONESIAN_MONTHS.items():
            if v_str == m_name.upper():
                month_col_map[m_num] = c
                break

    # Scan baris 4-7 untuk mapping kolom metadata (baris terbawah prioritas tertinggi jika ada tumpang tindih)
    for row_idx in [4, 5, 6, 7]:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row_idx, c).value
            if v is None:
                continue
            v_str = str(v).strip().upper()
            if v_str in ('NO', 'NO.') and cols['no'] is None:
                cols['no'] = c
            elif 'FLT' in v_str and 'NO' in v_str:
                cols['flight'] = c
            elif v_str == 'TO':
                cols['to'] = c
            elif v_str == 'ETD' or v_str == 'STD':
                cols['etd'] = c
            elif v_str == 'ETA' or v_str == 'STA':
                cols['eta'] = c
            elif v_str == 'ATD':
                cols['atd'] = c
            elif v_str == 'ATA':
                cols['ata'] = c
            elif 'PERIODE' in v_str:
                cols['periode'] = c
            elif 'SURAT' in v_str:
                cols['surat'] = c
            elif 'TIPE' in v_str or 'PENGAJUAN' in v_str:
                cols['tipe'] = c
            elif 'DAY' in v_str or 'HARI' in v_str:
                cols['hari'] = c
            elif 'TOTAL' in v_str or 'REALISASI' in v_str:
                cols['total'] = c

    # Kolom hari pertama: cari kolom dengan header angka "1" di baris 6 atau 7
    for row_idx in [6, 7]:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row_idx, c).value
            if str(v).strip() == '1' or v == 1:
                cols['day_start'] = c
                break
        if cols.get('day_start'):
            break

    if cols['no'] is None:
        cols['no'] = 1

    return cols, month_col_map


# ---------------------------------------------------------------------------
# Find Flight Headers
# ---------------------------------------------------------------------------

def _find_flight_headers(ws, col_flight):
    """
    Temukan semua baris awal 'blok flight' di kolom B (flight).
    Return list of (row_num, flight_str_normalized).
    Setiap blok flight mencakup beberapa baris (satu per bulan musim).
    """
    headers = []
    last_flight = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col_flight).value
        if v is None:
            last_flight = None
            continue
        v_str = str(v).strip()
        norm = _normalize_flight(v_str)
        # Hanya ambil baris pertama dari masing-masing flight number
        if re.match(r'^QG\d+[A-Z]?$', norm):
            if norm != last_flight:
                headers.append((r, norm, v_str))
                last_flight = norm
        else:
            last_flight = None
    return headers


# ---------------------------------------------------------------------------
# Block Extent
# ---------------------------------------------------------------------------

def _block_end(ws, start, next_start, total_rows):
    """Tentukan baris akhir dari sebuah blok flight."""
    if next_start is not None:
        return next_start - 1
    # Blok terakhir: cari baris kosong pertama (tidak ada data di col 13 / bulan_tahun atau col 2)
    # Biasanya setelah blok flight ada 1-2 baris kosong sebelum Keterangan/Signature
    for r in range(start + 1, total_rows + 1):
        has_data = False
        for c in range(2, 14):
            if ws.cell(r, c).value is not None and str(ws.cell(r, c).value).strip() != '':
                has_data = True
                break
        
        if not has_data:
            return r - 1
            
    return total_rows


# ---------------------------------------------------------------------------
# Copy Row
# ---------------------------------------------------------------------------

def _copy_row(ws, src_row, dst_row):
    """Salin nilai, style, dan tinggi baris dari src ke dst, termasuk menerjemahkan formula."""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        
        val = src.value
        if isinstance(val, str) and val.startswith('='):
            try:
                dst.value = Translator(val, src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = val
        else:
            dst.value = val
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


# ---------------------------------------------------------------------------
# Copy Merged Ranges for Block
# ---------------------------------------------------------------------------

def _copy_merged_ranges_for_block(ws, src_start, dst_start, block_len):
    """Duplikasi seluruh merged ranges yang sepenuhnya berada di dalam blok."""
    to_add = []
    # Kumpulkan range yang akan diduplikasi (menggunakan list untuk menghindari mutasi saat iterasi)
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= src_start and rng.max_row < src_start + block_len:
            offset = dst_start - src_start
            to_add.append((rng.min_row + offset, rng.min_col,
                           rng.max_row + offset, rng.max_col))
    
    for sr, sc, er, ec in to_add:
        # Hindari menduplikasi jika range sudah ada (walau kemungkinannya kecil)
        overlap = False
        for ex_rng in ws.merged_cells.ranges:
            if ex_rng.min_row <= sr and ex_rng.max_row >= er and ex_rng.min_col <= sc and ex_rng.max_col >= ec:
                overlap = True
                break
        if not overlap:
            ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)


# ---------------------------------------------------------------------------
# Shift Merged Ranges
# ---------------------------------------------------------------------------

def _shift_merged_ranges_below(ws, row_idx, amount):
    """Geser koordinat merged cell di bawah row_idx sebanyak amount baris."""
    to_remove = []
    to_add = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= row_idx:
            to_remove.append(str(rng))
            to_add.append((rng.min_row + amount, rng.min_col,
                           rng.max_row + amount, rng.max_col))
        elif rng.min_row < row_idx <= rng.max_row:
            to_remove.append(str(rng))
            if rng.min_row < row_idx - 1:
                to_add.append((rng.min_row, rng.min_col,
                               row_idx - 1, rng.max_col))
            to_add.append((row_idx + amount, rng.min_col,
                           rng.max_row + amount, rng.max_col))
    for rng_str in to_remove:
        # Cari dan hapus merged range berdasarkan string (kompatibel semua versi openpyxl)
        for existing_rng in list(ws.merged_cells.ranges):
            if str(existing_rng) == rng_str:
                ws.merged_cells.ranges.discard(existing_rng)
                break
    for sr, sc, er, ec in to_add:
        if sr <= er:
            ws.merge_cells(start_row=sr, start_column=sc,
                           end_row=er, end_column=ec)


# ---------------------------------------------------------------------------
# Shift Formulas
# ---------------------------------------------------------------------------

def _shift_formulas_below(ws, row_idx, amount):
    """Update seluruh formula di bawah row_idx karena ada penyisipan baris."""
    for r in range(row_idx + amount, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                old_coord = ws.cell(r - amount, c).coordinate
                try:
                    cell.value = Translator(val, old_coord).translate_formula(cell.coordinate)
                except Exception:
                    pass


# ---------------------------------------------------------------------------
# Rebuild Totals
# ---------------------------------------------------------------------------

def _rebuild_totals(ws, start, end, cols, month_col_map):
    """Perbaiki rumus SUM/COUNTIF di kolom Total Realisasi."""
    if not cols.get('total'):
        return
    for r in range(start, end + 1):
        cell = ws.cell(r, cols['total'])
        cell_type = type(cell).__name__
        if cell_type == 'MergedCell':
            continue
        if not cell.value and r != start:
            continue
        # Hitung rentang kolom harian untuk baris ini
        if month_col_map:
            all_day_cols = []
            for m_num in sorted(month_col_map.keys()):
                m_start = month_col_map[m_num]
                days = calendar.monthrange(2026, m_num)[1]
                all_day_cols.extend(range(m_start, m_start + days))
            if all_day_cols:
                c_start = get_column_letter(min(all_day_cols))
                c_end = get_column_letter(max(all_day_cols))
                cell.value = f'=SUM({c_start}{r}:{c_end}{r})'


# ---------------------------------------------------------------------------
# Main: Generate Report
# ---------------------------------------------------------------------------

def generate_report(project_id: int, template_path: str, output_path: str) -> int:
    """
    Fungsi utama pembuatan laporan.
    
    Args:
        project_id: ID project dari database.
        template_path: Path ke file template Excel (jika kosong, cari dari project.template_path).
        output_path: Path output file Excel.
    
    Returns:
        Jumlah blok penerbangan yang diisi.
    """
    # --- Ambil data dari database ---
    project = Project.objects.get(id=project_id)
    year = project.year
    month = project.month

    # Tentukan template path
    if not template_path:
        template_path = project.template_path
        
    # Fallback ke static template
    if not template_path or not os.path.exists(template_path):
        from django.conf import settings
        static_template = os.path.join(settings.BASE_DIR, 'static', 'tpl', 'form_realisasi_winter26.xlsx')
        if os.path.exists(static_template):
            template_path = static_template
        else:
            raise FileNotFoundError(
                f"Template Excel tidak ditemukan: '{template_path}' maupun di folder static. "
                "Pastikan file form_realisasi_winter26.xlsx ada di dalam folder static/tpl/."
            )

    # --- Ambil semua schedule versions untuk project ini ---
    schedules = ScheduleVersion.objects.filter(
        project_id=project_id,
    ).order_by('flight_number', 'version_number', 'flight_date')

    # Build struktur data:
    # flight_data[flight_norm] = {
    #   'wtt': {flight_str, origin, destination, std, sta, atd, ata},
    #   'pprp': [{periode, pprp_letter, std, sta, atd, ata}],  # jika ada
    #   'daily': {(year, month, day): operational_flag}         # dari GHP
    # }
    flight_data = {}
    for sv in schedules:
        fn = _normalize_flight(sv.flight_number)
        if fn not in flight_data:
            flight_data[fn] = {
                'flight_str': sv.flight_number,
                'origin': sv.origin,
                'destination': sv.destination,
                'wtt': None,
                'wtt_start_date': None,   # tanggal pertama WTT (untuk periode SEMULA)
                'wtt_end_date': None,     # tanggal terakhir WTT (untuk periode MENJADI)
                'pprp_list': [],
                'daily': {},  # (year, month, day) -> 1 or 0
            }

        d = sv.flight_date
        # Jangan timpa angka 1 dengan 0 jika ada duplikasi versi jadwal
        if sv.operational_flag:
            flight_data[fn]['daily'][(d.year, d.month, d.day)] = 1
        elif (d.year, d.month, d.day) not in flight_data[fn]['daily']:
            flight_data[fn]['daily'][(d.year, d.month, d.day)] = 0

        if sv.version_number == 1 and not sv.pprp_letter:
            # Ini baris WTT asli — lacak tanggal awal dan akhir musim
            fd = flight_data[fn]
            if fd['wtt_start_date'] is None or sv.flight_date < fd['wtt_start_date']:
                fd['wtt_start_date'] = sv.flight_date
            if fd['wtt_end_date'] is None or sv.flight_date > fd['wtt_end_date']:
                fd['wtt_end_date'] = sv.flight_date
            # Set metadata WTT (hanya sekali, dari record pertama)
            if not fd['wtt']:
                fd['wtt'] = {
                    'std': sv.std.strftime('%H:%M') if sv.std else None,
                    'sta': sv.sta.strftime('%H:%M') if sv.sta else None,
                    'atd': sv.atd.strftime('%H:%M') if sv.atd else None,
                    'ata': sv.ata.strftime('%H:%M') if sv.ata else None,
                }
        elif sv.pprp_letter:
            # Baris PPRP — satu record per surat PPRP per flight
            existing_letters = [p['pprp_letter'] for p in flight_data[fn]['pprp_list']]
            if sv.pprp_letter not in existing_letters:
                # Periode MENJADI: pprp_date s/d akhir musim (wtt_end_date)
                # wtt_end_date mungkin belum terisi saat ini (tergantung urutan query),
                # akan di-resolve setelah loop selesai
                flight_data[fn]['pprp_list'].append({
                    'pprp_letter': sv.pprp_letter,
                    'pprp_date': sv.pprp_date,          # tanggal mulai berlaku PPRP
                    'periode': '',                       # akan dihitung setelah loop
                    'std': sv.std.strftime('%H:%M') if sv.std else None,
                    'sta': sv.sta.strftime('%H:%M') if sv.sta else None,
                    'atd': sv.atd.strftime('%H:%M') if sv.atd else None,
                    'ata': sv.ata.strftime('%H:%M') if sv.ata else None,
                })

    # --- Hitung string periode setelah loop selesai (wtt_end_date sudah terisi) ---
    for fn, fd in flight_data.items():
        wtt_start = fd.get('wtt_start_date')
        wtt_end = fd.get('wtt_end_date')
        for pprp in fd['pprp_list']:
            d_pprp = pprp.get('pprp_date')
            if not d_pprp:
                continue
            import datetime as _dt
            semula_end = d_pprp - _dt.timedelta(days=1)
            # Simpan info periode di dict pprp untuk dipakai saat inject template
            if wtt_start:
                pprp['periode_semula'] = (
                    f"{wtt_start.day} {MONTH_ABBR[wtt_start.month]} {wtt_start.year}"
                    f"/{semula_end.day} {MONTH_ABBR[semula_end.month]} {semula_end.year}"
                )
            else:
                pprp['periode_semula'] = ''
            if wtt_end:
                pprp['periode'] = (
                    f"{d_pprp.day} {MONTH_ABBR[d_pprp.month]} {d_pprp.year}"
                    f"/{wtt_end.day} {MONTH_ABBR[wtt_end.month]} {wtt_end.year}"
                )

    # --- Buka template Excel ---
    wb = load_workbook(template_path)
    ws = wb.active

    # Deteksi kolom
    cols, month_col_map = _detect_columns(ws)

    col_flight = cols['flight']

    # Temukan semua blok flight di template
    headers = _find_flight_headers(ws, col_flight)
    if not headers:
        raise RuntimeError(
            "Tidak ada blok flight ditemukan di template. "
            "Pastikan kolom B berisi nomor penerbangan QG-xxx."
        )

    # --- Proses duplikasi PPRP (dari bawah ke atas agar baris tidak bergeser) ---
    for i in reversed(range(len(headers))):
        row_start, flight_norm, flight_str = headers[i]
        fd = flight_data.get(flight_norm)
        if not fd:
            continue

        # --- UPDATE ORIGINAL BLOCK (WTT/SEMULA) ---
        # Isi ATD, ATA asli dari data WTT (ETD/ETA biarkan bawaan template)
        if fd.get('wtt'):
            if cols.get('atd') and fd['wtt'].get('atd'):
                ws.cell(row_start, cols['atd']).value = fd['wtt']['atd']
            if cols.get('ata') and fd['wtt'].get('ata'):
                ws.cell(row_start, cols['ata']).value = fd['wtt']['ata']

        # Update SEMULA block: pastikan Tipe Pengajuan diset ke 'Perpanjangan'
        if cols.get('tipe'):
            target_row = row_start
            for rng in ws.merged_cells.ranges:
                if (rng.min_col <= cols['tipe'] <= rng.max_col and
                        rng.min_row <= row_start <= rng.max_row):
                    target_row = rng.min_row
                    break
            cell = ws.cell(target_row, cols['tipe'])
            if type(cell).__name__ != 'MergedCell':
                cell.value = 'Perpanjangan'

        # Jika tidak ada PPRP, skip proses duplikasi blok
        if not fd['pprp_list']:
            continue

        # Tentukan batas blok
        next_row = headers[i + 1][0] if i + 1 < len(headers) else None
        row_end = _block_end(ws, row_start, next_row, ws.max_row)
        block_len = row_end - row_start + 1

        for pprp in reversed(fd['pprp_list']):
            # --- UPDATE SEMULA block: ubah periode end date menjadi pprp_date - 1 ---
            if cols.get('periode') and pprp.get('periode_semula'):
                periode_col = cols['periode']
                # Cari top-left cell dari merged range di kolom periode pada baris start
                # (merged cell hanya bisa ditulis melalui top-left cell-nya)
                target_row = row_start
                for rng in ws.merged_cells.ranges:
                    if (rng.min_col <= periode_col <= rng.max_col and
                            rng.min_row <= row_start <= rng.max_row):
                        target_row = rng.min_row
                        break
                cell = ws.cell(target_row, periode_col)
                if type(cell).__name__ != 'MergedCell':
                    cell.value = pprp['periode_semula']

            # --- INSERT MENJADI block setelah blok SEMULA ---
            insert_at = row_end + 1
            ws.insert_rows(insert_at, block_len)
            _shift_merged_ranges_below(ws, insert_at, block_len)
            _shift_formulas_below(ws, insert_at, block_len)

            # Salin seluruh blok SEMULA ke posisi MENJADI
            for offset in range(block_len):
                _copy_row(ws, row_start + offset, insert_at + offset)
                
            # Salin struktur merged cells dari blok SEMULA ke MENJADI
            _copy_merged_ranges_for_block(ws, row_start, insert_at, block_len)

            # Update nilai di baris pertama blok MENJADI
            top_row = insert_at
            # Kosongkan kolom NO (akan di-rebuild)
            ws.cell(top_row, cols['no']).value = None
            # Update Periode MENJADI
            if cols.get('periode') and pprp['periode']:
                ws.cell(top_row, cols['periode']).value = pprp['periode']
            # Update Nomor Surat PPRP
            if cols.get('surat') and pprp['pprp_letter']:
                ws.cell(top_row, cols['surat']).value = pprp['pprp_letter']
            # Update Tipe Pengajuan MENJADI
            if cols.get('tipe'):
                ws.cell(top_row, cols['tipe']).value = 'Perubahan'
            # Update jam (ETD/ETA/ATD/ATA) dari PPRP
            if cols.get('etd') and pprp['std']:
                ws.cell(top_row, cols['etd']).value = pprp['std']
            if cols.get('eta') and pprp['sta']:
                ws.cell(top_row, cols['eta']).value = pprp['sta']
            if cols.get('atd') and pprp['atd']:
                ws.cell(top_row, cols['atd']).value = pprp['atd']
            if cols.get('ata') and pprp['ata']:
                ws.cell(top_row, cols['ata']).value = pprp['ata']
            # Kosongkan kolom harian (akan diisi ulang berdasarkan GHP)
            if cols.get('day_start') and month_col_map:
                for m_num, m_col_start in month_col_map.items():
                    days_in_month = calendar.monthrange(year, m_num)[1]
                    for offset in range(block_len):
                        for day_col in range(m_col_start, m_col_start + days_in_month):
                            ws.cell(insert_at + offset, day_col).value = None

            row_end = insert_at + block_len - 1  # Update row_end

    # --- Refresh headers setelah insert ---
    headers = _find_flight_headers(ws, col_flight)

    # --- Isi data harian (1/0) per blok per bulan ---
    for i, (row_start, flight_norm, flight_str) in enumerate(headers):
        fd = flight_data.get(flight_norm)
        if not fd or not fd['daily']:
            continue

        next_row = headers[i + 1][0] if i + 1 < len(headers) else None
        row_end = _block_end(ws, row_start, next_row, ws.max_row)

        day_start_col = cols.get('day_start')
        bulan_tahun_col = cols.get('bulan_tahun', 13)
        if not day_start_col:
            continue

        # Iterasi setiap baris di dalam blok flight ini
        for r in range(row_start, row_end + 1):
            bln_val = ws.cell(r, bulan_tahun_col).value
            if not bln_val:
                continue
                
            m_num = None
            if isinstance(bln_val, (datetime.datetime, datetime.date)):
                m_num = bln_val.month
            else:
                bln_str = str(bln_val).strip().lower()
                for num, name in INDONESIAN_MONTHS.items():
                    if name.lower()[:3] in bln_str or MONTH_ABBR[num].lower() in bln_str:
                        m_num = num
                        break
            
            if not m_num:
                continue
                
            # Isi dari tanggal 1 sampai akhir bulan
            days_in_month = calendar.monthrange(year, m_num)[1]
            for day in range(1, days_in_month + 1):
                day_col = day_start_col + (day - 1)
                flag = fd['daily'].get((year, m_num, day), 0)
                cell = ws.cell(r, day_col)
                if type(cell).__name__ != 'MergedCell':
                    cell.value = flag

    # --- Rebuild nomor urut (kolom A) ---
    seq = 1
    for row_start, flight_norm, flight_str in _find_flight_headers(ws, col_flight):
        cell = ws.cell(row_start, cols['no'])
        cell_type = type(cell).__name__
        if cell_type != 'MergedCell':
            cell.value = seq
        seq += 1

    # --- Update Tanggal Pengesahan (Current Date) ---
    now = datetime.datetime.now()
    tgl_sekarang = f"Surabaya, {INDONESIAN_MONTHS.get(now.month, 'Januari')} {now.year}"
    
    # Cari sel signature (mulai dari bawah untuk efisiensi)
    for r in range(ws.max_row, max(1, ws.max_row - 100), -1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val and isinstance(val, str) and 'Surabaya' in val:
                ws.cell(r, c).value = tgl_sekarang

    # --- Simpan output ---
    wb.save(output_path)
    return seq - 1  # Jumlah blok yang diisi


def load_template_flight_metadata(template_path: str = None) -> dict:
    """
    Ekstrak metadata baseline per flight dari template Excel resmi:
    - route (misal: 'SUB-CGK', 'SUB-BTH', 'SUB-BDJ')
    - periode (misal: '29 MAR 2026/24 OKT 2026')
    - surat (misal: 'AU.012/29/19/DRJU-DAU-2026')
    - tipe (misal: 'Perpanjangan')
    - start_date & end_date (parsed datetime.date)
    """
    MONTHS_ID = {'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MEI': 5, 'JUN': 6, 'JUL': 7, 'AGU': 8, 'SEP': 9, 'OKT': 10, 'NOV': 11, 'DES': 12}
    
    def _parse_d(s):
        m = re.search(r'(\d{1,2})\s+([A-Z]{3})\s+(\d{4})', s.upper())
        if m:
            d = int(m.group(1))
            mon = MONTHS_ID.get(m.group(2), 1)
            yr = int(m.group(3))
            return datetime.date(yr, mon, d)
        return None

    path_to_use = template_path if (template_path and os.path.exists(template_path)) else os.path.join('static', 'tpl', 'form_realisasi_winter26.xlsx')
    if not os.path.exists(path_to_use):
        return {}

    try:
        wb = load_workbook(path_to_use, data_only=True)
        ws = wb.active
        meta = {}
        for r in range(8, ws.max_row + 1):
            flt = ws.cell(r, 2).value
            if flt and re.match(r'^QG[- ]?\d+', str(flt).strip(), re.IGNORECASE):
                norm = _normalize_flight(flt)
                route = ws.cell(r, 3).value
                periode = ws.cell(r, 8).value
                surat = ws.cell(r, 9).value
                tipe = ws.cell(r, 10).value
                
                start_d, end_d = None, None
                if periode and '/' in str(periode):
                    parts = str(periode).split('/')
                    start_d = _parse_d(parts[0])
                    end_d = _parse_d(parts[1]) if len(parts) > 1 else None
                    
                meta[norm] = {
                    'route': str(route).strip() if route else '',
                    'periode': str(periode).strip() if periode else '',
                    'surat': str(surat).strip() if surat else '',
                    'tipe': str(tipe).strip() if tipe else 'Perpanjangan',
                    'start_date': start_d,
                    'end_date': end_d,
                }
        return meta
    except Exception:
        return {}


def format_id_date_abbr(d: datetime.date) -> str:
    """Format tanggal ke '29 MAR 2026'."""
    if not d:
        return ''
    m_abbr = MONTH_ABBR.get(d.month, '')
    return f"{d.day:02d} {m_abbr} {d.year}"


def get_project_report_data(project_id: int) -> dict:
    """
    Mengambil dan menyusun data rekonsiliasi penerbangan dari database
    untuk ditampilkan pada Report Viewer HTML maupun diekspor ke Excel.
    Hanya mengambil penerbangan keberangkatan dari Surabaya (origin='SUB').
    """
    project = Project.objects.get(id=project_id)
    tpl_meta = load_template_flight_metadata(project.template_path)
    
    all_schedules = ScheduleVersion.objects.filter(
        project=project,
        origin='SUB'
    ).order_by('flight_number', 'version_number', 'flight_date')
    
    from collections import defaultdict
    flight_groups = defaultdict(list)
    for s in all_schedules:
        flight_groups[s.flight_number].append(s)
        
    rows = []
    grand_planned = 0
    grand_operated = 0
    
    periode_label = f"Summer {str(project.year)[-2:]} (S-{str(project.year)[-2:]})" if project.month in range(3, 11) else f"Winter {str(project.year)[-2:]} (W-{str(project.year)[-2:]})"
    month_abbr_name = INDONESIAN_MONTHS.get(project.month, '')
    month_label = f"{month_abbr_name[:3]}-{str(project.year)[-2:]}"
    
    row_seq = 1
    for f_num in sorted(flight_groups.keys()):
        items = flight_groups[f_num]
        v1_items = [s for s in items if s.version_number == 1]
        v2_items = [s for s in items if s.version_number == 2]
        
        f_norm = _normalize_flight(f_num)
        base_meta = tpl_meta.get(f_norm, {})
        
        # Route: SUB-XXX
        dest_code = v1_items[0].destination if v1_items else (v2_items[0].destination if v2_items else '')
        route_str = base_meta.get('route') or f"SUB-{dest_code}"
        
        # Base metadata defaults
        base_surat = base_meta.get('surat') or '-'
        base_tipe = 'Perpanjangan'
        base_periode = base_meta.get('periode') or f"29 MAR {project.year}/24 OKT {project.year}"
        start_season_d = base_meta.get('start_date') or datetime.date(project.year, 3, 29)
        end_season_d = base_meta.get('end_date') or datetime.date(project.year, 10, 24)
        
        if not v2_items:
            # Case 1: Pure Baseline Flight (No PPRP changes)
            first_item = v1_items[0]
            days_dict = {s.flight_date.day: s for s in v1_items}
            
            total_planned = len(v1_items)
            total_operated = sum(1 for s in v1_items if s.operational_flag)
            grand_planned += total_planned
            grand_operated += total_operated
            
            pct = (total_operated / total_planned * 100) if total_planned > 0 else 0
            
            weekdays_present = {item.flight_date.isoweekday() for item in v1_items}
            pattern = "".join(str(d) if d in weekdays_present else "-" for d in range(1, 8))
            
            days_row = []
            for day in range(1, 32):
                if day in days_dict:
                    s = days_dict[day]
                    days_row.append({
                        'day': day,
                        'is_scheduled': True,
                        'is_operated': bool(s.operational_flag),
                        'val': '1' if s.operational_flag else '0'
                    })
                else:
                    days_row.append({
                        'day': day,
                        'is_scheduled': False,
                        'is_operated': False,
                        'val': '-'
                    })
                    
            rows.append({
                'no': row_seq,
                'flight_number': f_num,
                'origin': first_item.origin or 'SUB',
                'to': route_str,
                'etd': first_item.std.strftime('%H:%M') if first_item.std else '',
                'eta': first_item.sta.strftime('%H:%M') if first_item.sta else '',
                'atd': first_item.atd.strftime('%H:%M') if first_item.atd else (first_item.std.strftime('%H:%M') if first_item.std else ''),
                'ata': first_item.ata.strftime('%H:%M') if first_item.ata else (first_item.sta.strftime('%H:%M') if first_item.sta else ''),
                'periode': base_periode,
                'pprp_no': base_surat,
                'pprp_type': base_tipe,
                'day_pattern': pattern,
                'total_planned': total_planned,
                'month_label': month_label,
                'days': days_row,
                'total_operated': total_operated,
                'pct': round(pct, 1),
            })
            row_seq += 1
        else:
            # Case 2: Flight has PPRP changes (generate Row 1: SEMULA, Row 2: Perubahan)
            pprp_first = v2_items[0]
            pprp_start_date = pprp_first.pprp_date or v2_items[0].flight_date
            semula_end_date = pprp_start_date - datetime.timedelta(days=1)
            
            # Formatted period strings
            periode_semula = f"{format_id_date_abbr(start_season_d)}/{format_id_date_abbr(semula_end_date)}"
            periode_perubahan = f"{format_id_date_abbr(pprp_start_date)}/{format_id_date_abbr(end_season_d)}"
            
            # Determine submission_type from PPRP file if possible
            sub_type = 'Perubahan'
            if pprp_first.source_pprp:
                try:
                    from core.parsers.pprp import parse_pprp
                    sub_type = parse_pprp(pprp_first.source_pprp.file_path).get('submission_type', 'Perubahan')
                except Exception:
                    sub_type = 'Perubahan'
            
            # --- 1. Row SEMULA (WTT Baseline before PPRP start date) ---
            semula_items = [s for s in v1_items if s.flight_date < pprp_start_date]
            first_semula = v1_items[0] if v1_items else pprp_first
            semula_days_dict = {s.flight_date.day: s for s in semula_items}
            
            total_planned_semula = len(semula_items)
            total_operated_semula = sum(1 for s in semula_items if s.operational_flag)
            grand_planned += total_planned_semula
            grand_operated += total_operated_semula
            pct_semula = (total_operated_semula / total_planned_semula * 100) if total_planned_semula > 0 else 0
            
            weekdays_present_semula = {item.flight_date.isoweekday() for item in v1_items}
            pattern_semula = "".join(str(d) if d in weekdays_present_semula else "-" for d in range(1, 8))
            
            days_row_semula = []
            for day in range(1, 32):
                if day in semula_days_dict:
                    s = semula_days_dict[day]
                    days_row_semula.append({
                        'day': day,
                        'is_scheduled': True,
                        'is_operated': bool(s.operational_flag),
                        'val': '1' if s.operational_flag else '0'
                    })
                else:
                    days_row_semula.append({
                        'day': day,
                        'is_scheduled': False,
                        'is_operated': False,
                        'val': '-'
                    })
                    
            rows.append({
                'no': row_seq,
                'flight_number': f_num,
                'origin': first_semula.origin or 'SUB',
                'to': route_str,
                'etd': first_semula.std.strftime('%H:%M') if first_semula.std else '',
                'eta': first_semula.sta.strftime('%H:%M') if first_semula.sta else '',
                'atd': first_semula.atd.strftime('%H:%M') if first_semula.atd else (first_semula.std.strftime('%H:%M') if first_semula.std else ''),
                'ata': first_semula.ata.strftime('%H:%M') if first_semula.ata else (first_semula.sta.strftime('%H:%M') if first_semula.sta else ''),
                'periode': periode_semula,
                'pprp_no': base_surat,
                'pprp_type': base_tipe,
                'day_pattern': pattern_semula,
                'total_planned': total_planned_semula,
                'month_label': month_label,
                'days': days_row_semula,
                'total_operated': total_operated_semula,
                'pct': round(pct_semula, 1),
            })
            row_seq += 1
            
            # --- 2. Row MENJADI / PPRP (on and after PPRP start date) ---
            pprp_days_dict = {s.flight_date.day: s for s in v2_items}
            total_planned_pprp = len(v2_items)
            total_operated_pprp = sum(1 for s in v2_items if s.operational_flag)
            grand_planned += total_planned_pprp
            grand_operated += total_operated_pprp
            pct_pprp = (total_operated_pprp / total_planned_pprp * 100) if total_planned_pprp > 0 else 0
            
            weekdays_present_pprp = {item.flight_date.isoweekday() for item in v2_items}
            pattern_pprp = "".join(str(d) if d in weekdays_present_pprp else "-" for d in range(1, 8))
            
            days_row_pprp = []
            for day in range(1, 32):
                if day in pprp_days_dict:
                    s = pprp_days_dict[day]
                    days_row_pprp.append({
                        'day': day,
                        'is_scheduled': True,
                        'is_operated': bool(s.operational_flag),
                        'val': '1' if s.operational_flag else '0'
                    })
                else:
                    days_row_pprp.append({
                        'day': day,
                        'is_scheduled': False,
                        'is_operated': False,
                        'val': '-'
                    })
                    
            rows.append({
                'no': row_seq,
                'flight_number': f_num,
                'origin': pprp_first.origin or 'SUB',
                'to': route_str,
                'etd': pprp_first.std.strftime('%H:%M') if pprp_first.std else '',
                'eta': pprp_first.sta.strftime('%H:%M') if pprp_first.sta else '',
                'atd': pprp_first.atd.strftime('%H:%M') if pprp_first.atd else (pprp_first.std.strftime('%H:%M') if pprp_first.std else ''),
                'ata': pprp_first.ata.strftime('%H:%M') if pprp_first.ata else (pprp_first.sta.strftime('%H:%M') if pprp_first.sta else ''),
                'periode': periode_perubahan,
                'pprp_no': pprp_first.pprp_letter or base_surat,
                'pprp_type': sub_type,
                'day_pattern': pattern_pprp,
                'total_planned': total_planned_pprp,
                'month_label': month_label,
                'days': days_row_pprp,
                'total_operated': total_operated_pprp,
                'pct': round(pct_pprp, 1),
            })
            row_seq += 1
        
    grand_pct = (grand_operated / grand_planned * 100) if grand_planned > 0 else 0
    now = datetime.datetime.now()
    signature_date = f"Surabaya, {now.day} {INDONESIAN_MONTHS.get(now.month, 'Januari')} {now.year}"
    
    return {
        'project': project,
        'periode_label': periode_label,
        'month_name': INDONESIAN_MONTHS.get(project.month, ''),
        'year': project.year,
        'rows': rows,
        'grand_planned': grand_planned,
        'grand_operated': grand_operated,
        'grand_pct': round(grand_pct, 1),
        'days_header': list(range(1, 32)),
        'signature_date': signature_date,
        'total_flight_groups': len(rows),
    }

